from openpyxl import load_workbook

wb = load_workbook('For-write.xlsx')
sheet = wb.active

# Write Column name
columns = sheet['A1'].value = "Name"
columns = sheet['B1'].value = "Age"
columns = sheet['C1'].value = "Salary"
print('Successfully write Column Name')

# insert data in columns
name = ['Muntasir', 'John', 'Smith']
sheet.cell(row=1, column=1).value = "Name"

j = 2
for i in range(0,3):
    sheet.cell(row=j, column=1).value = name[i]
    j +=1


age = ['18', '20', '25']
sheet.cell(row=1, column=2).value = "Age"

k = 2 #k is changeable for more write
for p in range(0,3): #p is changeable for more write
    sheet.cell(row=k, column=2).value = age[p]
    k +=1

wb.save('For-write.xlsx')
print('Successfully write Column Data')
